# utils/detect_utils.py
import re
import io
import cv2
import numpy as np
from .ocr_utils import pil_to_cv

# -------------------------------
# 📌 NER via spaCy
# -------------------------------
import spacy
_nlp = None

def get_spacy_model():
    global _nlp
    if _nlp is None:
        _nlp = spacy.load("en_core_web_sm")
    return _nlp

def ner_entities(text):
    nlp = get_spacy_model()
    doc = nlp(text)
    findings = []
    for ent in doc.ents:
        if ent.label_ in ["PERSON", "GPE", "LOC", "ORG"]:
            findings.append({
                "label": ent.label_,
                "span": [ent.start_char, ent.end_char],
                "matched": ent.text
            })
    return findings

# -------------------------------
# 📌 Regex patterns
# -------------------------------
PATTERNS = {
    "EMAIL": re.compile(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"),
    "PHONE": re.compile(r"(?<!\d)(?:\+91[-\s]?|0)?[6-9]\d{9}(?!\d)"),
    "DOB": re.compile(
        r"\b(?:DOB|Date of Birth|Birth Date)[:\s]*\d{1,2}[\/\-]\d{1,2}[\/\-](?:19|20)\d{2}\b", re.I),
    "AADHAAR": re.compile(r"(?<!\d)(?:\d{4}[- ]?\d{4}[- ]?\d{4})(?!\d)"),
    "PAN": re.compile(r"\b[A-Z]{5}[0-9]{4}[A-Z]\b"),
    # Add other patterns if needed
}

SIGNATURE_HINTS = ["signature", "sign"]

# -------------------------------
# 📌 Regex entity detection
# -------------------------------
def regex_entities(text):
    findings = []
    for label, pat in PATTERNS.items():
        for m in pat.finditer(text):
            findings.append({"label": label, "span": [m.start(), m.end()], "matched": m.group(0)})
    return findings

# -------------------------------
# 📌 Merge regex + NER
# -------------------------------
def get_all_text_findings(text):
    return regex_entities(text) + ner_entities(text)

# -------------------------------
# 📌 Map regex matches back to OCR lines
# -------------------------------
def align_text_findings_to_boxes(text_findings, ocr_lines):
    boxes = []
    for tf in text_findings:
        target = tf["matched"].strip()
        placed = False
        for line in ocr_lines:
            if target in line.get("text",""):
                boxes.append({"label": tf["label"], "box": line["box"], "matched": target})
                placed = True
                break
        if not placed:
            boxes.append({"label": tf["label"], "box": [0,0,0,0], "matched": target})
    return boxes

# -------------------------------
# 📌 Face detection
# -------------------------------
_face_cascade = None
def get_face_cascade():
    global _face_cascade
    if _face_cascade is None:
        _face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    return _face_cascade

def detect_faces(cv_img):
    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
    faces = get_face_cascade().detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4, minSize=(40,40))
    return [{"type":"FACE","box":[int(x),int(y),int(w),int(h)]} for (x,y,w,h) in faces]

# -------------------------------
# 📌 Signature detection
# -------------------------------
def find_signature_regions(cv_img, ocr_lines):
    h, w = cv_img.shape[:2]
    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5,5), 0)
    edges = cv2.Canny(blur, 50, 150)
    dil = cv2.dilate(edges, np.ones((3,3), np.uint8), iterations=1)
    regions = []
    for line in ocr_lines:
        text = line.get("text","").lower()
        if any(k in text for k in SIGNATURE_HINTS):
            x,y,w0,h0 = line["box"]
            y1 = min(h, y+h0+int(3*h0))
            x0 = max(0, x-int(0.2*w0))
            x1 = min(w, x+w0+int(0.2*w0))
            roi = dil[y:y1, x0:x1]
            cnts,_ = cv2.findContours(roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            for c in cnts:
                area = cv2.contourArea(c)
                if 500 < area < 40000:
                    rx, ry, rw, rh = cv2.boundingRect(c)
                    regions.append({"type":"SIGNATURE","box":[int(x0+rx), int(y+ry), int(rw), int(rh)]})
    return regions

# -------------------------------
# 📌 Excel helper
# -------------------------------
def mask_excel_cells_and_render(file_stream):
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_stream.read()), data_only=True)
    results = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        detections = []
        for r in ws.iter_rows(values_only=True):
            rows.append([("" if c is None else str(c)) for c in r])
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                if not val: continue
                s = str(val)
                for lab, pat in PATTERNS.items():
                    if pat.search(s):
                        detections.append({"label":lab,"value":s,"cell":f"{i+1},{j+1}","box":[40+j*200,80+i*24,180,22]})
                        rows[i][j] = "█"*min(len(s),20)
                        break
        img = Image.new("RGB",(1654,2339),"white")
        draw = ImageDraw.Draw(img)
        try: font = ImageFont.truetype("DejaVuSans.ttf",16)
        except: font = ImageFont.load_default()
        y=40
        for r in rows:
            line = " | ".join(r)
            draw.text((40,y),line[:2000],fill="black",font=font)
            y+=22
            if y>2280: break
        results.append((sheet,img,detections))
    return results
