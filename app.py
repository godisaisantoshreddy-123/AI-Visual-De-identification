# app.py
import os
import io
import json
from datetime import datetime

import streamlit as st
from PIL import Image, ImageDraw, ImageFont

import spacy

nlp = spacy.load("en_core_web_sm")

from utils.ocr_utils import get_ocr_reader, pil_to_cv, cv_to_pil, ocr_with_boxes
from utils.detect_utils import (
    regex_entities,
    align_text_findings_to_boxes,
    detect_faces,
    find_signature_regions,
    detect_qr_regions,
    ner_findings,
)
from utils.redact_utils import apply_redactions_cv
from utils.pdf_utils import load_pages_from_pdf, remove_pdf_metadata_bytes, export_images_to_pdf

# ---------- Folders ----------
OUTPUT_DIR = "outputs"
RED_PDFS = os.path.join(OUTPUT_DIR, "redacted_pdfs")
RED_LOGS = os.path.join(OUTPUT_DIR, "logs")
for d in (RED_PDFS, RED_LOGS):
    os.makedirs(d, exist_ok=True)

# ---------- Page ----------
st.set_page_config(page_title="AI Visual De-ID", layout="wide")
st.title("🔒 AI Visual De-identification – Privacy by Design")

with st.expander("Instructions"):
    st.markdown("""
    - Upload **images, PDFs, DOCX, XLSX** (multi-file supported).
    - Choose **Auto-Redact** to skip the detection table (faster & cleaner).
    - You’ll see a **progress bar**. When 100% is reached, downloads appear.
    - We redact by **Blur** or **Black box**, including **QR codes**, **faces**, **signatures**.
    - Extra PII: **SSN, NPI, Insurance IDs, ICD-10, Address heuristic**.
    - Optional: **spaCy NER** to catch names/addresses/orgs.
    """)

# ---------- Options ----------
colA, colB, colC, colD, colE = st.columns([1.5, 1.2, 1.2, 1.5, 1.5])
with colA:
    redact_method = st.radio("Redaction method", ["blur", "black"], horizontal=True)
with colB:
    fast_mode = st.checkbox("⚡ Fast OCR mode (downscale)", value=True)
with colC:
    enable_ner = st.checkbox("NER for names/addresses (spaCy)", value=True)
with colD:
    auto_redact = st.checkbox("Auto-redact (skip detection table)", value=True)
with colE:
    preview_mode = st.checkbox("⚡ High-speed Preview (no blur until export)", value=True)

colF, colG, colH = st.columns([1, 1, 1])
with colF:
    enable_faces = st.checkbox("Detect Faces", value=True)
with colG:
    enable_signatures = st.checkbox("Detect Signatures", value=True)
with colH:
    use_easyocr = st.checkbox("Use EasyOCR", value=True)

uploads = st.file_uploader(
    "Upload files (images, pdf, docx, xlsx). You can select multiple.",
    accept_multiple_files=True,
    type=["png", "jpg", "jpeg", "pdf", "docx", "xlsx"],
)

process = st.button("🚀 Process & Redact")

# ---------- State ----------
if "bundle" not in st.session_state:
    st.session_state.bundle = None
if "done_units" not in st.session_state:
    st.session_state.done_units = 0

def timestamp():
    return datetime.now().strftime("%Y%m%dT%H%M%S")

# ---------- Helpers ----------
@st.cache_resource
def _get_reader(use_easyocr: bool):
    return get_ocr_reader(use_easyocr)

def _prep_pil_image(file_obj):
    pil = Image.open(file_obj).convert("RGB")
    max_side = 1600 if fast_mode else 2800
    w, h = pil.size
    if max(w, h) > max_side:
        s = max_side / max(w, h)
        pil = pil.resize((int(w * s), int(h * s)), Image.LANCZOS)
    return pil

def process_one_pil(pil, name, reader, progress_cb, page_tag=""):
    cv_img = pil_to_cv(pil)
    lines = ocr_with_boxes(pil, reader)
    full_text = "\n".join([l["text"] for l in lines])

    text_findings = regex_entities(full_text)
    text_boxes = align_text_findings_to_boxes(text_findings, lines)

    ner_boxes = ner_findings(full_text, lines, enable_ner)

    face_boxes = detect_faces(cv_img) if enable_faces else []
    sig_boxes = find_signature_regions(cv_img, lines) if enable_signatures else []
    qr_boxes = detect_qr_regions(cv_img)  # ✅ QR detection included always

    merged = text_boxes + ner_boxes + face_boxes + sig_boxes + qr_boxes

    if auto_redact:
        selected_boxes = merged
    else:
        st.subheader(f"Detections {page_tag}".strip())
        cols_hdr = st.columns([3, 3, 4, 2])
        cols_hdr[0].markdown("**Type**")
        cols_hdr[1].markdown("**Box**")
        cols_hdr[2].markdown("**Sample**")
        cols_hdr[3].markdown("**Redact?**")
        selected_boxes = []
        for i, det in enumerate(merged):
            label = det.get("label", det.get("type", "PII"))
            box = det.get("box", [0, 0, 0, 0])
            sample = (det.get("matched") or det.get("text") or "")[:60]
            c0, c1, c2, c3 = st.columns([3, 3, 4, 2])
            c0.write(label)
            c1.write(f"{box[0]},{box[1]} ({box[2]}×{box[3]})")
            c2.write(sample)
            do = c3.checkbox(
                f"{label}: {sample}" if sample else label,
                value=True,
                key=f"{name}{page_tag}_redact_{i}",
                label_visibility="collapsed",
            )
            if do:
                selected_boxes.append(det)

    # ✅ High-speed preview mode
    if preview_mode:
        preview_img = pil.copy()
        draw = ImageDraw.Draw(preview_img)
        for det in selected_boxes:
            x, y, w, h = det.get("box", [0, 0, 0, 0])
            draw.rectangle([x, y, x + w, y + h], outline="red", width=3)
        st.image(preview_img, caption=f"Preview (fast mode) {page_tag}".strip(), use_container_width=True)
        redacted_pil = pil  # store original for later export
    else:
        redacted_pil = apply_redactions_cv(cv_img, selected_boxes, mode=redact_method)
        st.image(redacted_pil, caption=f"Redacted preview {page_tag}".strip(), use_container_width=True)

    progress_cb()
    return redacted_pil, {
        "file": name,
        "detections": merged,
        "redacted": selected_boxes,
        "mode": redact_method,
        "timestamp": timestamp(),
        "page": page_tag,
    }

# ---------- Main ----------
if process and uploads:
    reader = _get_reader(use_easyocr)

    total_units = 0
    for f in uploads:
        ext = os.path.splitext(f.name)[1].lower()
        if ext == ".pdf":
            raw = f.read(); f.seek(0)
            try:
                pages = load_pages_from_pdf(remove_pdf_metadata_bytes(raw))
                total_units += len(pages)
            except Exception:
                total_units += 1
        else:
            total_units += 1

    st.session_state.done_units = 0
    prog = st.progress(0.0)

    def progress_cb():
        st.session_state.done_units += 1
        prog.progress(min(1.0, st.session_state.done_units / max(1, total_units)))

    all_logs = []
    redacted_images_global = []

    for file in uploads:
        st.markdown(f"### 📄 Processing {file.name}")
        name = file.name
        ext = os.path.splitext(name)[1].lower()

        if ext in [".png", ".jpg", ".jpeg"]:
            pil = _prep_pil_image(file)
            red_pil, log = process_one_pil(pil, name, reader, progress_cb)
            redacted_images_global.append(red_pil); all_logs.append(log)

        elif ext == ".pdf":
            raw = file.read(); file.seek(0)
            stripped = remove_pdf_metadata_bytes(raw)
            pages = load_pages_from_pdf(stripped)
            for pno, pil in enumerate(pages, start=1):
                red_pil, log = process_one_pil(pil, name, reader, progress_cb, page_tag=f"(p{pno})")
                redacted_images_global.append(red_pil); all_logs.append(log)

        elif ext == ".docx":
            from utils.doc_utils import extract_text_from_docx_stream
            text = extract_text_from_docx_stream(file)
            img = Image.new("RGB", (1654, 2339), "white")
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("DejaVuSans.ttf", 18)
            except:
                font = ImageFont.load_default()
            y = 40
            for line in text.splitlines():
                draw.text((40, y), line[:1800], fill="black", font=font)
                y += 24
                if y > 2300:
                    break
            pil = img
            red_pil, log = process_one_pil(pil, name, reader, progress_cb)
            redacted_images_global.append(red_pil); all_logs.append(log)

        elif ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
                img = Image.new("RGB", (1654, 2339), "white")
                draw = ImageDraw.Draw(img)
                try:
                    font = ImageFont.truetype("DejaVuSans.ttf", 16)
                except:
                    font = ImageFont.load_default()
                y = 40
                for r in rows:
                    line = " | ".join(r)
                    draw.text((40, y), line[:2000], fill="black", font=font)
                    y += 22
                    if y > 2280:
                        break
                red_pil, log = process_one_pil(img, f"{name} [{sheet}]", reader, progress_cb)
                redacted_images_global.append(red_pil); all_logs.append(log)

        else:
            st.warning(f"Unsupported file type: {ext}")
            progress_cb()

    if redacted_images_global:
        # ✅ If preview mode, now apply real redaction for export
        if preview_mode:
            real_redacted_images = []
            for log, pil in zip(all_logs, redacted_images_global):
                cv_img = pil_to_cv(pil)
                red_pil = apply_redactions_cv(cv_img, log["redacted"], mode=redact_method)
                real_redacted_images.append(red_pil)
        else:
            real_redacted_images = redacted_images_global

        pdf_bytes = export_images_to_pdf(real_redacted_images).getvalue()
        ts = timestamp()
        pdf_name = f"redacted_{ts}.pdf"
        log_name = f"redaction_log_{ts}.json"

        st.session_state.bundle = {
            "pdf_bytes": pdf_bytes,
            "pdf_name": pdf_name,
            "log_json": json.dumps(all_logs, indent=2).encode("utf-8"),
            "log_name": log_name,
        }

        with open(os.path.join(RED_PDFS, pdf_name), "wb") as f:
            f.write(pdf_bytes)
        with open(os.path.join(RED_LOGS, log_name), "wb") as f:
            f.write(st.session_state.bundle["log_json"])

        st.success("✅ All done — files are ready below.")
    else:
        st.info("No redacted output was produced.")
        st.session_state.bundle = None

# ---------- Download section ----------
if st.session_state.bundle:
    st.download_button(
        "⬇️ Download Redacted PDF",
        data=st.session_state.bundle["pdf_bytes"],
        file_name=st.session_state.bundle["pdf_name"],
        mime="application/pdf",
        key="download_pdf",
    )
    st.download_button(
        "⬇️ Download JSON Redaction Log",
        data=st.session_state.bundle["log_json"],
        file_name=st.session_state.bundle["log_name"],
        mime="application/json",
        key="download_log",
    )
